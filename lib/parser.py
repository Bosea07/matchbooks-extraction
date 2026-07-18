"""
MatchBooks file extraction - generic heuristic column detection.

Ported from the tested browser prototype (matchbooks-reconciliation-engine.html),
which used the same header-synonym matching to turn arbitrary Excel/CSV SOA
exports into normalized {ref, date, type, amount} records.

Scope note: this is Layer 2 of the planned "universal SOA format recognition"
pipeline (generic heuristic column detection). Known-template signature
matching, confidence scoring, and the Claude API fallback for genuinely
unrecognizable formats are separate, larger additions for later.
"""

import csv
import io
import math
import re
from datetime import date, datetime

DATE_KEYS = ['date', 'txn date', 'transaction date', 'invoice date', 'bill date', 'posting date']
REF_KEYS = [
    'ref', 'reference', 'ref no', 'ref#', 'reference no', 'reference number',
    'invoice no', 'invoice #', 'invoice number', 'bill no', 'bill number',
    'voucher no', 'voucher', 'document no', 'doc no'
]
TYPE_KEYS = ['type', 'transaction type', 'txn type', 'doc type']
DEBIT_KEYS = ['debit', 'debit amount', 'dr', 'dr amount']
CREDIT_KEYS = ['credit', 'credit amount', 'cr', 'cr amount']
AMOUNT_KEYS = ['amount', 'invoice amount', 'total amount', 'value', 'net amount']


def _is_blank(v):
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return str(v).strip() == ''


def normalize_header(h):
    return re.sub(r'\s+', ' ', ('' if h is None else str(h)).strip().lower())


def find_col(headers, keys):
    norm = [normalize_header(h) for h in headers]
    for k in keys:
        for idx, h in enumerate(norm):
            if h == k:
                return idx
    for k in keys:
        for idx, h in enumerate(norm):
            if k in h:
                return idx
    return -1


def parse_amount(v):
    if _is_blank(v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == '' or s == '-':
        return 0.0
    neg = False
    if re.match(r'^\(.*\)$', s):
        neg = True
        s = s[1:-1]
    s = re.sub(r'[^0-9.\-]', '', s)
    try:
        n = float(s)
    except ValueError:
        return 0.0
    return -abs(n) if neg else n


def normalize_ref(ref):
    return re.sub(r'\s+', '', ('' if ref is None else str(ref)).strip().upper())


def format_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime('%d-%m-%Y')
    if _is_blank(v):
        return ''
    return str(v)


def sheet_to_records(rows):
    """rows: list of lists (raw grid, as read from CSV/XLSX)."""
    if not rows:
        return {'records': [], 'meta': {}}

    def non_empty_count(row):
        return sum(1 for c in row if not _is_blank(c))

    header_row_idx = next((i for i, r in enumerate(rows) if non_empty_count(r) >= 2), 0)
    headers = rows[header_row_idx]

    date_idx = find_col(headers, DATE_KEYS)
    ref_idx = find_col(headers, REF_KEYS)
    type_idx = find_col(headers, TYPE_KEYS)
    debit_idx = find_col(headers, DEBIT_KEYS)
    credit_idx = find_col(headers, CREDIT_KEYS)
    amount_idx = find_col(headers, AMOUNT_KEYS)

    def cell(r, idx):
        if idx == -1 or idx >= len(r):
            return None
        return r[idx]

    records = []
    for i in range(header_row_idx + 1, len(rows)):
        r = rows[i]
        if r is None or all(_is_blank(c) for c in r):
            continue

        ref_raw = cell(r, ref_idx)
        ref = normalize_ref(ref_raw)
        if not ref:
            continue

        if debit_idx != -1 or credit_idx != -1:
            d = parse_amount(cell(r, debit_idx))
            c = parse_amount(cell(r, credit_idx))
            amount = d - c
        elif amount_idx != -1:
            amount = parse_amount(cell(r, amount_idx))
        else:
            amount = 0.0

        records.append({
            'ref': ref,
            'refRaw': '' if ref_raw is None else str(ref_raw),
            'date': format_date(cell(r, date_idx)),
            'type': '' if _is_blank(cell(r, type_idx)) else str(cell(r, type_idx)),
            'amount': amount
        })

    return {
        'records': records,
        'meta': {
            'headerRow': header_row_idx,
            'totalRows': len(rows),
            'dateCol': date_idx,
            'refCol': ref_idx,
            'typeCol': type_idx,
            'debitCol': debit_idx,
            'creditCol': credit_idx,
            'amountCol': amount_idx,
        }
    }


def _read_csv_rows(content: bytes):
    text = None
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError('Could not decode CSV file (tried utf-8, latin-1).')
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def _read_excel_rows(content: bytes):
    try:
        import openpyxl
    except ImportError as err:
        raise ValueError('openpyxl is not installed on the server.') from err

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def extract_from_bytes(filename: str, content: bytes):
    ext = filename.lower().rsplit('.', 1)[-1] if filename and '.' in filename else ''

    if ext == 'csv':
        rows = _read_csv_rows(content)
    elif ext in ('xlsx', 'xlsm'):
        rows = _read_excel_rows(content)
    elif ext == 'xls':
        raise ValueError(
            'Legacy .xls format is not supported yet - please save as .xlsx or .csv and re-upload.'
        )
    else:
        raise ValueError(f'Unsupported file type: .{ext or "unknown"}. Supported: .xlsx, .csv')

    return sheet_to_records(rows)
